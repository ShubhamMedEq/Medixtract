import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import os

# Import extractors and utility functions
from extractors.pdf1_extractor import extract_billing_info_pdf1
from extractors.pdf2_extractor import extract_billing_info_pdf2
from extractors.pdf3_extractor import extract_billing_info_pdf3
from extractors.utils import read_excel

def process_pdfs(uploaded_files):
    """Extract data from PDFs and return a merged DataFrame."""
    data_frames = []

    for uploaded_file in uploaded_files:
        file_name = uploaded_file.name

        # Save the uploaded file temporarily
        with open(file_name, 'wb') as f:
            f.write(uploaded_file.getbuffer())

        # Use appropriate extractor based on the file name
        if "R000345" in file_name:
            df = extract_billing_info_pdf1(file_name)
        elif "R000548" in file_name:
            df = extract_billing_info_pdf2(file_name)
        elif "R000530" in file_name:
            df = extract_billing_info_pdf3(file_name)
        else:
            st.warning(f"Unsupported PDF: {file_name}")
            continue

        data_frames.append(df)
        os.remove(file_name)  # Clean up temp file

    if data_frames:
        return pd.concat(data_frames, ignore_index=True)
    else:
        return pd.DataFrame()

def update_excel_sheet(data, excel_file_path):
    """Update the ANALYSIS sheet in the Excel file."""
    workbook = load_workbook(excel_file_path)
    if 'ANALYSIS' in workbook.sheetnames:
        df_analysis = pd.read_excel(excel_file_path, sheet_name='ANALYSIS', header=None)
    else:
        st.error("ANALYSIS sheet does not exist.")
        return None

    # Identify the header row
    existing_columns = None
    for i in range(len(df_analysis)):
        if df_analysis.iloc[i].str.upper().isin(
            ["PROVIDER", "DATE", "CPT", "DESCRIPTION", "CHARGES"]
        ).any():
            existing_columns = df_analysis.iloc[i].str.upper().tolist()
            break

    if existing_columns is None:
        st.error("No valid header row found.")
        return None

    # Map columns to their indices
    col_indices = {col: existing_columns.index(col) if col in existing_columns else None
                   for col in ["PROVIDER", "DATE", "CPT", "DESCRIPTION", "CHARGES"]}

    # Add PROVIDED_DESCRIPTION column if missing
    if "PROVIDED_DESCRIPTION" not in existing_columns:
        provided_description_index = len(existing_columns)
        df_analysis.loc[0, provided_description_index] = 'PROVIDED_DESCRIPTION'
        df_analysis[provided_description_index] = None
        existing_columns.append('PROVIDED_DESCRIPTION')
    else:
        provided_description_index = existing_columns.index("PROVIDED_DESCRIPTION")

    # Insert data into the Excel DataFrame
    for _, row in data.iterrows():
        new_row = [None] * len(existing_columns)

        if col_indices["PROVIDER"] is not None:
            new_row[col_indices["PROVIDER"]] = row['PROVIDER']
        if col_indices["DATE"] is not None:
            new_row[col_indices["DATE"]] = row['DATE'].strftime('%m/%d/%Y')
        if col_indices["CPT"] is not None:
            new_row[col_indices["CPT"]] = row['CPT']
        if col_indices["DESCRIPTION"] is not None:
            new_row[col_indices["DESCRIPTION"]] = row['DESCRIPTION']
        if col_indices["CHARGES"] is not None:
            new_row[col_indices["CHARGES"]] = row['CHARGES']

        new_row[provided_description_index] = row['PROVIDED_DESCRIPTION']
        df_analysis.loc[len(df_analysis)] = new_row

    # Save the updated workbook
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df_analysis.to_excel(writer, sheet_name='ANALYSIS', index=False, header=False)

    return df_analysis

# Streamlit App Layout
st.title("PDF Billing Data Extractor and Excel Updater")

uploaded_files = st.file_uploader("Upload PDF files", type=["pdf"], accept_multiple_files=True)
uploaded_excel = st.file_uploader("Upload Excel file", type=["xlsx"])

if st.button("Process and Update Excel"):
    if uploaded_files and uploaded_excel:
        # Save the uploaded Excel file locally
        excel_path = uploaded_excel.name
        with open(excel_path, 'wb') as f:
            f.write(uploaded_excel.getbuffer())

        # Extract data from PDFs
        merged_data = process_pdfs(uploaded_files)

        if not merged_data.empty:
            st.success("PDF data extracted successfully!")
            st.write(merged_data)

            # Read Excel CPT data and merge with extracted PDF data
            excel_data = read_excel(excel_path)
            final_data = pd.merge(merged_data, excel_data, left_on='CPT', right_on='Procedure Code', how='left')
            final_data.drop(columns=['Procedure Code'], inplace=True)
            final_data.rename(columns={'Description': 'PROVIDED_DESCRIPTION'}, inplace=True)

            # Update the Excel sheet
            updated_df = update_excel_sheet(final_data, excel_path)

            if updated_df is not None:
                st.success(f"Excel file '{excel_path}' updated successfully!")
                st.write(updated_df)

                # Create a downloadable CSV from the final data
                csv_data = final_data.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Download Output CSV",
                    data=csv_data,
                    file_name="output.csv",
                    mime="text/csv"
                )
            else:
                st.error("Failed to update the Excel sheet.")
        else:
            st.warning("No data extracted from PDFs. Please check the files.")
    else:
        st.warning("Please upload both PDF files and an Excel file.")
